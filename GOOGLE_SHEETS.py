# Указание кодировки для файла, это важно для корректной работы с не ASCII символами
# -*- coding: utf-8 -*-

import pandas as pd  # Библиотека для работы с таблицами данных
import gspread  # Библиотека для работы с Google Sheets
from oauth2client.service_account import ServiceAccountCredentials  # Инструменты для авторизации через Google API
from gspread_dataframe import set_with_dataframe  # Функция для записи DataFrame в Google Sheets
from gspread_formatting import *  # Функции для форматирования ячеек в Google Sheets
from openpyxl import load_workbook  # Для работы с файлами Excel
import logging  # Для логирования

# Настройка логирования: указываем файл для записи, уровень логирования и формат сообщений
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler('changes.log', mode='w'),  # Логи будут записываться в файл changes.log
                        logging.StreamHandler()  # Логи также будут выводиться в стандартный поток вывода
                    ]
                )
try:
    # Настройка доступа к Google Sheets API с помощью файла учетных данных
    scopes = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    creds = ServiceAccountCredentials.from_json_keyfile_name('path_to_credentials.json', scopes) #в нужно вставить название своего файла с доступом к google
    client = gspread.authorize(creds)

    # Открываем Google Sheets документ по названию и выбираем первый лист
    sheet = client.open('table_for_work').sheet1
    urls = sheet.col_values(6)  # Получаем значения из шестого столбца

    # Загружаем существующий Excel файл и выбираем активный лист
    wb = load_workbook('companies.xlsx')
    ws = wb.active

    # Очищаем значения в 13-м столбце (столбец L)
    for row in ws.iter_rows(min_col=13, max_col=13):
        for cell in row:
            cell.value = None

    # Записываем URL в столбец L начиная с первой строки
    for idx, url in enumerate(urls, start=1):
        ws.cell(row=idx, column=13, value=url)

    # Сохраняем изменения в Excel файле
    wb.save('companies.xlsx')

    # Читаем данные из Excel файла в DataFrame
    df = pd.read_excel('companies.xlsx')

    # Заменяем пропущенные значения на пустые строки
    df.fillna("", inplace=True)

    # Приводим столбцы с датами и ИНН к строковому формату, чтобы избежать проблем с типами данных
    df['LAST_DATE'] = df['LAST_DATE'].astype('str')
    df['INN'] = df['INN'].astype('str')

    # Выбираем столбцы в нужном порядке для дальнейшей работы
    selected_columns = df[["№", "NAME_FOR_FOLDER", "INN", "STATUS", "LAST_DATE", "URL"]]

    # Очищаем все данные на листе Google Sheets перед записью новых
    sheet.clear()

    # Записываем данные из DataFrame в Google Sheets начиная с первой строки и столбца
    set_with_dataframe(sheet, selected_columns, row=1, col=1, include_column_header=True)

    # Форматируем ячейки с определённым статусом, красим в светло-красный цвет
    status_column_index = 5  # Статус находится в пятом столбце выбранных данных
    red_background_format = CellFormat(backgroundColor=Color(1, 0.8, 0.8))
    for index, row in selected_columns.iterrows():
        if row['STATUS'] != 'Действующая':
            format_cell_range(sheet, f'A{index + 2}:G{index + 2}', red_background_format)

    # Форматируем заголовки столбцов, делаем текст жирным
    format_cell_range(sheet, 'A1:E1', CellFormat(textFormat=TextFormat(bold=True)))

    # Добавляем границы вокруг ячеек
    format_cell_range(sheet, 'A1:E{}'.format(len(selected_columns) + 1), CellFormat(borders=Borders(top=Border('SOLID'), bottom=Border('SOLID'), left=Border('SOLID'), right=Border('SOLID'))))

    # Записываем информацию об успешном переносе данных в лог-файл
    logging.info('Данные успешно перенесены в Google sheets!')

except Exception as e:
    # В случае возникновения ошибки записываем информацию об ошибке в лог
    logging.error(f'Произошла ошибка: {str(e)}', exc_info=True)

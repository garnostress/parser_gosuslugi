import requests  # Импорт модуля для отправки HTTP-запросов
import openpyxl  # Импорт модуля для работы с файлами Excel формата .xlsx
from datetime import datetime  # Импорт класса datetime для работы с датами и временем
import logging  # Импорт модуля для логирования

# Настройка логирования для записи в файл 'changes.log' с уровнем важности INFO и определённым форматом сообщений
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler('changes.log', mode='w'),  # Логи будут записываться в файл changes.log
                        logging.StreamHandler()  # Логи также будут выводиться в стандартный поток вывода
                    ]
                )

API_TOKEN = ""  # Токен для доступа к API Яндекс.Диска
BASE_URL = "https://cloud-api.yandex.net/v1/disk/resources"  # Базовый URL для обращения к API Яндекс.Диска

# Функция для получения содержимого папки на Яндекс.Диске
def get_folder_content(url, token):
    headers = {"Authorization": f"OAuth {token}"}  # Заголовок для авторизации запроса с использованием токена
    response = requests.get(url, headers=headers)  # Отправка GET-запроса для получения содержимого папки
    return response.json()  # Возврат ответа в формате JSON

# Функция для получения дат из имен папок на Яндекс.Диске
def get_dates_from_folder(token, folder_path):
    url = f"{BASE_URL}?path={folder_path}"  # Формирование URL для запроса содержимого папки
    folder_content = get_folder_content(url, token)  # Получение содержимого папки

    if 'error' in folder_content:
        return None, None, False  # Возврат значений None, если в ответе есть ошибка

    if '_embedded' not in folder_content:
        return None, None, True  # Возврат значений None, если в ответе нет ключа '_embedded'

    all_dates = []  # Список для хранения всех дат
    for item in folder_content['_embedded']['items']:  # Перебор всех элементов в папке
        if item['type'] == 'dir':  # Если элемент является папкой
            try:
                date_str = item['name'].split(" на")[0]  # Попытка извлечь дату из имени папки
                date = datetime.strptime(date_str, "%d.%m.%y")  # Преобразование строки в объект datetime
                all_dates.append(date)  # Добавление даты в список
            except ValueError:
                # Запись в лог предупреждения, если дата не может быть распознана
                logging.warning(f"Could not parse date from folder name: {item['name']}")

    all_dates.sort(reverse=True)  # Сортировка дат в обратном порядке (от новых к старым)
    # Возврат самой последней и предпоследней даты, если они есть
    return all_dates[0] if all_dates else 'Отсутствует', all_dates[1] if len(all_dates) > 1 else 'Отсутствует', True

# Загрузка книги Excel и выбор активного листа
wb = openpyxl.load_workbook("companies.xlsx")
sheet = wb.active

changed_rows = []  # Список для отслеживания изменённых строк

# Перебор строк в таблице, начиная со второй строки и до последней
for row_num, row in enumerate(sheet.iter_rows(min_row=2, max_col=11, max_row=sheet.max_row), start=2):
    number_cell = row[0]  # Первая ячейка строки содержит номер
    inn_cell = row[2]  # Третья ячейка строки содержит ИНН
    name_for_folder_cell = row[8]  # Девятая ячейка строки содержит имя для папки
    name_gosuslugi_cell = row[1]  # Вторая ячейка строки содержит имя, зарегистрированное в Госуслугах
    name_for_folder = name_for_folder_cell.value  # Значение из ячейки имени для папки
    inn = int(inn_cell.value)  # Значение ИНН, преобразованное в целое число
    number = number_cell.value  # Значение номера
    name_gosuslugi = name_gosuslugi_cell.value  # Значение имени, зарегистрированного в Госуслугах

    folder_name = f"{number}. {name_for_folder} ({inn})"  # Формирование имени папки
    folder_path = f"disk:/ИПР/{folder_name}"  # Формирование пути к папке на Яндекс.Диске

    # Получение последней и предпоследней даты обновления папки
    latest_date, before_latest_date, folder_exists = get_dates_from_folder(API_TOKEN, folder_path)

    # Если папка не существует, запись предупреждения в лог
    if not folder_exists:
        logging.warning(f"Folder does not exist for row number {row_num}")
        continue

    # Сохранение старых значений для последующего сравнения
    old_val_10 = sheet.cell(row=row_num, column=10).value
    old_val_11 = sheet.cell(row=row_num, column=11).value

    # Запись новых дат в лист Excel
    sheet.cell(row=row_num, column=10).value = latest_date.strftime("%d.%m.%Y") if isinstance(latest_date, datetime) else 'Отсутствует'
    sheet.cell(row=row_num, column=11).value = before_latest_date.strftime("%d.%m.%Y") if isinstance(before_latest_date, datetime) else 'Отсутствует'

    # Получение новых значений для сравнения с старыми
    new_val_10 = sheet.cell(row=row_num, column=10).value
    new_val_11 = sheet.cell(row=row_num, column=11).value

    # Если значения изменились, запись информации об изменении в лог
    if old_val_10 != new_val_10 or old_val_11 != new_val_11:
        logging.info(f"Row {row_num}: Changes detected for {name_for_folder} ({inn}) - "
                     f"last update date: {old_val_10} -> {new_val_10}, "
                     f"before last update date: {old_val_11} -> {new_val_11}")

# Сохранение изменений в файл Excel
wb.save("companies.xlsx")

# Запись информации о сохранении изменений в файл журнала
logging.info("Changes saved to companies.xlsx")

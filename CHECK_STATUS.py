import openpyxl # Импорт библиотеки для работы с Excel файлами
from openpyxl.styles import PatternFill # Импорт класса для применения заливки ячеек в Excel
from datetime import datetime # Импорт класса для работы с датой и временем
from dadata import Dadata # Импорт клиента DaData для работы с данными компаний
import pandas as pd # Импорт библиотеки для анализа и обработки данных
import time # Импорт модуля для работы с системным временем
import logging # Импорт модуля для логирования

# Настройка базовых параметров логирования: имя файла, уровень логирования и формат сообщений
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler('changes.log', mode='w'),  # Логи будут записываться в файл changes.log
                        logging.StreamHandler()  # Логи также будут выводиться в стандартный поток вывода
                    ]
                )

# Определение функции для извлечения имени компании из текста
def extract_name_in_quotes(text, gosuslugi_name):
    # Проверка на NaN (не число) значения
    if pd.isna(text):
        return text

    # Список ключевых слов для проверки в тексте
    keywords = ["ТРАНСЭНЕРГО", "ГАЗПРОМ ЭНЕРГОСБЫТ ТЮМЕНЬ", "ОБОРОНЭНЕРГО", "ДЭК", "ОЭК", "АТОМЭНЕРГОСБЫТ"]

    # Преобразование текста в верхний регистр для унификации сравнения
    text = str(text).upper()

    # Проверка на наличие ключевых слов в тексте или извлечение текста в кавычках
    if any(keyword in text for keyword in keywords):
        return gosuslugi_name
    elif '"' in text:
        start = text.find('"') + 1  # Индекс начала текста в кавычках
        end = text.find('"', start)  # Индекс конца текста в кавычках
        return text[start:end]  # Возврат текста между кавычками
    else:
        return text  # Возврат исходного текста, если условия не выполнены

# Основной цикл выполнения
while True:
    try:
        # Токен для аутентификации в API DaData
        token = "" #В кавычки нужно вставить API ключ Dadata
        # Инициализация клиента DaData с токеном
        dadata = Dadata(token)

        # Словарь для перевода статусов компаний
        status_translation = {
            "ACTIVE": "Действующая",
            "LIQUIDATING": "Ликвидируется",
            "LIQUIDATED": "Ликвидирована",
            "BANKRUPT": "Банкротство",
            "REORGANIZING": "В процессе присоединения к другому юрлицу, с последующей ликвидацией"
        }

        # Загрузка рабочей книги Excel
        workbook = openpyxl.load_workbook('companies.xlsx')
        # Получение активного листа рабочей книги
        sheet = workbook.active

        # Сбор ИНН из колонки C и текущих статусов из колонки D
        inn_list = [cell.value for cell in sheet['C'] if cell.value is not None]
        current_statuses = [cell.value for cell in sheet['D'] if cell.value is not None]

        # Инициализация списков для новых статусов, временных меток и дополнительных данных
        new_statuses = []
        timestamps = []
        values = []
        full_with_opfs = []

        # Инициализация списка для сбора данных
        data = []

        # Обработка каждой записи в списке ИНН и текущих статусов
        for index, (inn, current_status) in enumerate(zip(inn_list[1:], current_statuses[1:]), 2):
            # Поиск данных о компании по ИНН через API DaData
            result = dadata.find_by_id("party", inn)
            # Получение статуса компании из ответа API
            api_status = result[0]['data']['state']['status'] if result and 'data' in result[0] and 'state' in result[0][
                'data'] and 'status' in result[0]['data']['state'] else 'Неизвестно'
            # Перевод статуса на русский язык
            new_status = status_translation.get(api_status, 'Неизвестно')
            # Добавление нового статуса в список
            new_statuses.append(new_status)
            # Запись текущей временной метки
            timestamps.append(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

            # Получение краткого и полного наименования компании из ответа API
            value_data = result[0]['data']['name']['short_with_opf'] if result and 'data' in result[0] and 'name' in \
                                                                               result[0]['data'] and 'full_with_opf' in \
                                                                               result[0]['data']['name'] else 'Неизвестно'
            full_with_opf_data = result[0]['data']['name']['full_with_opf'] if result and 'data' in result[0] and 'name' in \
                                                                               result[0]['data'] and 'full_with_opf' in \
                                                                               result[0]['data']['name'] else 'Неизвестно'

            # Добавление полученных данных в соответствующие списки
            values.append(value_data)
            full_with_opfs.append(full_with_opf_data)

            # Проверка на изменение статуса и добавление данных в список, если есть изменения
            if new_status != current_status:
                data.append({
                    "GOSUSLUGI_NAME": sheet[f'B{index}'].value,
                    "OLD_STATUS": current_status,
                    "NEW_STATUS": new_status
                })

        # Создание DataFrame из списка изменений
        changes_df = pd.DataFrame(data)

        # Логирование изменений, если они есть
        if not changes_df.empty:
            logging.info('Изменения в компаниях:')
            for _, row in changes_df.iterrows():
                logging.info(f"{row['GOSUSLUGI_NAME']} изменил статус с {row['OLD_STATUS']} на {row['NEW_STATUS']}.")
        else:
            logging.info('Изменений нет.')

        # Обновление листа Excel данными о новых статусах, временных метках и названиях компаний
        for index, (status, timestamp, value, full_with_opf) in enumerate(zip(new_statuses, timestamps, values, full_with_opfs),
                                                                          start=2):
            sheet[f'D{index}'].value = status
            sheet[f'E{index}'].value = timestamp
            sheet[f'G{index}'].value = value
            sheet[f'H{index}'].value = full_with_opf
            sheet[f'I{index}'].value = extract_name_in_quotes(sheet[f'G{index}'].value, sheet[f'B{index}'].value)

        # Применение красной заливки для неактивных компаний
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        for index, status in enumerate(new_statuses, start=2):
            if status != "Действующая":
                for col_letter in ['A', 'B', 'C', 'E', 'D', 'F', 'G', 'H', 'I', 'J']:
                    sheet[f'{col_letter}{index}'].fill = red_fill

        # Сохранение изменений в файл Excel
        workbook.save('companies.xlsx')
        # Выход из цикла
        break

    except Exception as e:
        # Логирование исключения с трассировкой стека
        logging.error(f'Произошла ошибка: {str(e)}', exc_info=True)
        # Пауза перед повторной попыткой выполнения цикла
        time.sleep(5)
